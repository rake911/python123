from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename='regression corn database.xlsx', read_only=True)
print wb.get_sheet_names()
ws = wb["Sheet2"]

cell_range = ws['B1:k1300']
cell_range2 = ws['B1:k1']

names = {}
smplsize = 2

for row in cell_range2:
    for cell in row:
        names[cell.value] = []
print names

x = []
count = 0
count2 = 0 - smplsize + 1

for column in cell_range:
    count += 1
    if ws.columns[count][0].value in names:
        names[ws.columns[count][0].value] = ws.columns[count][:smplsize]
        for num in names[ws.columns[count][0].value]:
            num = num.value
            x.append(num)
            #print x
        count2 += smplsize
        names[ws.columns[count][0].value] = x[count2:]

print names

