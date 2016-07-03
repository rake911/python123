from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename = 'regression corn database.xlsx')
ws = wb.active

cell_range = ws['B1:k1']

names = {}

for row in cell_range:
    for cell in row:
        names[cell.value] = []
print names

corn = []
soybean = []

count = 0
for row in ws.iter_rows('B1:K1300'):

    y = row[0].value
    corn.append(y)
    if corn[0] in names:
        names[y] = corn

    z = row[1].value
    soybean.append(z)
    if soybean[0] in names:
        names[z] = soybean

for row in ws.iter_rows('B1:K1300'):
    count += 1
    try:
        print names[row[count].value]
    except:
        continue
print count
smplsize = 10
count2 = 0 - smplsize + 1
print count2