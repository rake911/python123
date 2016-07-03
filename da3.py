from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename='regression corn database.xlsx', read_only=True)
print wb.get_sheet_names()
ws = wb["Sheet2"]

cell_range = ws['B1:k1300']
cell_range2 = ws['B1:k1']

names = {}
smplsize = 100

for column in ws['B1:k10']:
    print column[0].value
    names[column[0].value] = []
print names